"""Evaluate a workbook validation contract's declared invariants against a saved workbook.

Per RFC 0002 decision 6, this module proves two things and nothing more:

* completion evidence -- did a required sheet/name/table actually get
  created, does a freshness marker look non-stale by an explicit rule, are
  there visible cached errors;
* declared invariants -- assertions a human (or a contract author) wrote
  down in advance: row-count minimums, sentinel cell values, freshness
  windows.

It does NOT attempt to prove the workbook's *data* is semantically correct
-- there is no ground truth available to this tool for that. "Every
declared invariant passed" means "the assertions a human wrote down in
advance held," never "the numbers are right." Nobody should read a passing
result from this module as proof of correctness.

Reads the workbook with openpyxl(data_only=True): the *cached* values as
last saved, not live formula recalculation. This is deliberate -- it is
what lets this module catch a stale cached workbook that shows no visible
formula errors, which is exactly the acceptance criterion this module
exists to satisfy ("a stale cached workbook cannot pass solely because
formula cells contain no visible errors").

Every assertion declared in the contract produces its own invariant_result
entry, even when an earlier assertion in the same contract already failed
(a missing required sheet does not abort evaluation of the rest of the
contract) -- a caller sees the full picture in one pass. Every entry
matches #34's `invariant_result` shape exactly: `{"name": str, "passed":
bool, "message": str (optional)}`, so these can be dropped straight into a
#34 result document's `invariants` array.
"""

from __future__ import annotations

import importlib.util
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
from openpyxl.utils.exceptions import InvalidFileException

from .errors import ContractError
from .schemas import validate_contract

_CHECK_FORMULA_ERRORS_PATH = (
    Path(__file__).resolve().parent.parent.parent / "scripts" / "check_formula_errors.py"
)

# Excel calculation modes that count as "automatic" for expected_calculation_mode.
# openpyxl exposes the raw OOXML calcPr@calcMode values ("auto", "autoNoTable",
# "manual"); the contract's friendly enum is just "automatic"/"manual".
_CALC_MODE_ALIASES = {
    "automatic": frozenset({"auto", "autoNoTable"}),
    "manual": frozenset({"manual"}),
}


def _load_excel_errors() -> frozenset:
    """Import EXCEL_ERRORS from xlsx-win/scripts/check_formula_errors.py, by file path.

    That script lives outside this package (xlsx-win/scripts, not
    xlsx-win/v2/control_plane), and "xlsx-win" contains a hyphen, so it can
    never be a dotted `-m` import target (see cli.py's own comment on this).
    Loading it by path -- rather than copying its EXCEL_ERRORS list -- means
    this module automatically tracks the repo's one canonical list of
    significant Excel error values instead of maintaining a second one that
    could drift out of sync.
    """
    spec = importlib.util.spec_from_file_location(
        "_xlsx_win_check_formula_errors", _CHECK_FORMULA_ERRORS_PATH
    )
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return frozenset(module.EXCEL_ERRORS)


EXCEL_ERRORS = _load_excel_errors()


def _open_workbook(workbook_path: Path):
    try:
        return load_workbook(workbook_path, data_only=True)
    except (InvalidFileException, KeyError, OSError, ValueError) as exc:
        raise ContractError(
            "WORKBOOK_UNREADABLE",
            f"Could not open workbook for validation: {exc}",
            {"workbook_path": str(workbook_path)},
        ) from exc


def _collect_tables(wb) -> dict:
    """Map table name -> (worksheet, Table) across every sheet in the workbook.

    ws.tables is an openpyxl TableList (a dict subclass) keyed by table
    name, but its own .items() is overridden to yield (name, ref-string)
    pairs rather than (name, Table). Indexing (ws.tables[name]) still
    returns the real Table object, so this iterates keys and indexes
    explicitly instead of using .items().
    """
    tables = {}
    for ws in wb.worksheets:
        for name in ws.tables.keys():
            tables[name] = (ws, ws.tables[name])
    return tables


def _result(name: str, passed: bool, message: str | None = None) -> dict:
    result = {"name": name, "passed": passed}
    if message is not None:
        result["message"] = message
    return result


def _check_required_sheets(wb, sheet_names: list) -> list:
    results = []
    for sheet_name in sheet_names:
        exists = sheet_name in wb.sheetnames
        message = None if exists else f"Sheet {sheet_name!r} does not exist in the workbook."
        results.append(_result(f"required_sheet:{sheet_name}", exists, message))
    return results


def _check_required_defined_names(wb, names: list) -> list:
    results = []
    for defined_name in names:
        exists = defined_name in wb.defined_names
        message = (
            None if exists else f"Defined name {defined_name!r} does not exist in the workbook."
        )
        results.append(_result(f"required_defined_name:{defined_name}", exists, message))
    return results


def _check_required_tables(table_names: list, tables: dict) -> list:
    results = []
    for table_name in table_names:
        exists = table_name in tables
        message = (
            None
            if exists
            else f"Table (ListObject) {table_name!r} does not exist in the workbook."
        )
        results.append(_result(f"required_table:{table_name}", exists, message))
    return results


def _sheet_row_count(ws) -> int:
    """Number of rows containing at least one non-empty cell, including any header row."""
    return sum(1 for row in ws.iter_rows(values_only=True) if any(v is not None for v in row))


def _table_row_count(table) -> int:
    """Number of data rows in a Table's range, excluding its own header row(s)."""
    min_col, min_row, max_col, max_row = range_boundaries(table.ref)
    header_rows = table.headerRowCount if table.headerRowCount is not None else 0
    return max(0, (max_row - min_row + 1) - header_rows)


def _check_min_row_counts(wb, min_row_counts: dict, tables: dict) -> list:
    results = []
    for name, minimum in min_row_counts.items():
        invariant_name = f"min_row_count:{name}"
        if name in tables:
            _, table = tables[name]
            actual = _table_row_count(table)
        elif name in wb.sheetnames:
            actual = _sheet_row_count(wb[name])
        else:
            results.append(
                _result(invariant_name, False, f"No sheet or table named {name!r} exists.")
            )
            continue

        passed = actual >= minimum
        message = None if passed else f"Expected at least {minimum} row(s), found {actual}."
        results.append(_result(invariant_name, passed, message))
    return results


def _read_cell(wb, sheet: str, cell: str):
    """Return (value, error_message). error_message is None on success."""
    if sheet not in wb.sheetnames:
        return None, f"Sheet {sheet!r} does not exist in the workbook."
    try:
        value = wb[sheet][cell].value
    except (ValueError, KeyError) as exc:
        return None, f"Could not read cell {sheet}!{cell}: {exc}"
    return value, None


def _check_sentinel_cells(wb, sentinel_cells: list) -> list:
    results = []
    for entry in sentinel_cells:
        sheet, cell, expected = entry["sheet"], entry["cell"], entry["expected"]
        invariant_name = f"sentinel_cell:{sheet}!{cell}"

        value, error = _read_cell(wb, sheet, cell)
        if error is not None:
            results.append(_result(invariant_name, False, error))
            continue

        passed = value == expected
        message = None if passed else f"Expected {expected!r}, found {value!r}."
        results.append(_result(invariant_name, passed, message))
    return results


def _as_datetime(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime(value.year, value.month, value.day)
    return None


def _check_freshness(wb, freshness: dict, now: datetime) -> list:
    sheet, cell, max_age_hours = freshness["sheet"], freshness["cell"], freshness["max_age_hours"]
    invariant_name = f"freshness:{sheet}!{cell}"

    value, error = _read_cell(wb, sheet, cell)
    if error is not None:
        return [_result(invariant_name, False, error)]

    as_of = _as_datetime(value)
    if as_of is None:
        return [
            _result(
                invariant_name,
                False,
                f"Cell {sheet}!{cell} does not contain a datetime/date value (found {value!r}).",
            )
        ]

    age_hours = (now - as_of).total_seconds() / 3600.0
    passed = age_hours <= max_age_hours
    message = (
        None
        if passed
        else (
            f"As-of timestamp {as_of.isoformat()} is {age_hours:.2f} hour(s) old; "
            f"maximum allowed age is {max_age_hours} hour(s)."
        )
    )
    return [_result(invariant_name, passed, message)]


def _check_prohibit_visible_errors(wb) -> list:
    locations = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "e" and cell.value in EXCEL_ERRORS:
                    locations.append(f"{ws.title}!{cell.coordinate} = {cell.value}")

    passed = not locations
    message = None
    if not passed:
        shown = locations[:10]
        suffix = "" if len(locations) <= 10 else f" (+{len(locations) - 10} more)"
        message = f"Found {len(locations)} visible error cell(s): {', '.join(shown)}{suffix}"
    return [_result("prohibit_visible_errors", passed, message)]


def _check_calculation_mode(wb, expected_mode: str) -> list:
    calc_props = getattr(wb, "calculation", None)
    calc_mode = getattr(calc_props, "calcMode", None) if calc_props is not None else None

    if calc_mode is None:
        return [
            _result(
                "expected_calculation_mode",
                False,
                "not_checkable: openpyxl could not read a calculation mode from this "
                "workbook's calcPr. Reporting failed-closed rather than silently "
                "passing or crashing.",
            )
        ]

    passed = calc_mode in _CALC_MODE_ALIASES[expected_mode]
    message = (
        None
        if passed
        else f"Expected calculation mode {expected_mode!r}, found {calc_mode!r}."
    )
    return [_result("expected_calculation_mode", passed, message)]


def evaluate_contract(workbook_path, contract: dict, *, now: datetime | None = None) -> list:
    """Evaluate every assertion in `contract` against the saved workbook at `workbook_path`.

    Returns a list of dicts, each matching #34's invariant_result shape
    exactly: {"name": str, "passed": bool, "message": str (optional)}.

    Raises ContractError if `contract` itself fails schema validation
    (SCHEMA_INVALID -- a caller bug, distinct from any individual invariant
    failing) or if the workbook cannot be opened at all (WORKBOOK_UNREADABLE).
    Every assertion declared in the contract gets its own entry in the
    returned list even if earlier assertions already failed.

    `now` is the evaluation time used for freshness checks; defaults to
    datetime.now(). Exposed as a keyword so callers (and tests) can pin it.
    """
    validate_contract(contract)
    workbook_path = Path(workbook_path)
    now = now if now is not None else datetime.now()

    wb = _open_workbook(workbook_path)
    try:
        tables = _collect_tables(wb)
        results: list = []

        results.extend(_check_required_sheets(wb, contract.get("required_sheets", [])))
        results.extend(
            _check_required_defined_names(wb, contract.get("required_defined_names", []))
        )
        results.extend(_check_required_tables(contract.get("required_tables", []), tables))
        results.extend(_check_min_row_counts(wb, contract.get("min_row_counts", {}), tables))
        results.extend(_check_sentinel_cells(wb, contract.get("sentinel_cells", [])))

        if "freshness" in contract:
            results.extend(_check_freshness(wb, contract["freshness"], now))

        if contract.get("prohibit_visible_errors", True):
            results.extend(_check_prohibit_visible_errors(wb))

        if "expected_calculation_mode" in contract:
            results.extend(_check_calculation_mode(wb, contract["expected_calculation_mode"]))

        return results
    finally:
        wb.close()
