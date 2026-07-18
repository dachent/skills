"""File-backend router tests: deterministic boolean routing (issue #35).

Covers the acceptance criteria carried over from RFC 0002's descope: a
plain xlsx routes to openpyxl; a macro-enabled xlsm, a signed workbook, and
a workbook with an external link all route to excel_required with the
triggering field(s) named in `explain`; a nonexistent xlsx/xlsm path with
create_new intent routes to xlsxwriter; an .xls path under edit_existing
routes to convert_required, while under create_new it fails closed instead
(there is nothing to convert-then-edit); .csv/.tsv route to not_applicable
under both create_new and edit_existing, never to xlsxwriter or openpyxl;
and a lightweight xlsxwriter-vs-openpyxl timing sanity check for the
new-workbook fast path.
"""

from __future__ import annotations

import time
from dataclasses import replace

import openpyxl
import pytest
import xlsxwriter

from control_plane.file_router import (
    BACKEND_CONVERT_REQUIRED,
    BACKEND_EXCEL_REQUIRED,
    BACKEND_NOT_APPLICABLE,
    BACKEND_OPENPYXL,
    BACKEND_XLSXWRITER,
    choose_backend,
)
from control_plane.workbook_inventory import WorkbookInventory

_BASE_INVENTORY = WorkbookInventory(
    path="C:\\jobs\\model.xlsx", exists=True, file_format="xlsx", sheet_count=1
)


def _inventory(**overrides) -> WorkbookInventory:
    return replace(_BASE_INVENTORY, **overrides)


def test_create_new_with_no_existing_file_routes_to_xlsxwriter() -> None:
    inventory = _inventory(exists=False)

    decision = choose_backend("create_new", inventory)

    assert decision.backend == BACKEND_XLSXWRITER
    assert decision.explain["exists"] is False


def test_create_new_over_an_existing_file_fails_closed() -> None:
    inventory = _inventory(exists=True)

    decision = choose_backend("create_new", inventory)

    assert decision.backend == BACKEND_EXCEL_REQUIRED


@pytest.mark.parametrize("file_format", ["csv", "tsv"])
@pytest.mark.parametrize("exists", [False, True])
def test_create_new_csv_or_tsv_is_declared_out_of_scope_not_xlsxwriter(
    file_format, exists
) -> None:
    inventory = _inventory(file_format=file_format, exists=exists)

    decision = choose_backend("create_new", inventory)

    assert decision.backend == BACKEND_NOT_APPLICABLE
    assert decision.backend != BACKEND_XLSXWRITER
    assert decision.explain.get("file_format") == file_format


@pytest.mark.parametrize("exists", [False, True])
def test_create_new_xls_fails_closed_not_xlsxwriter(exists) -> None:
    inventory = _inventory(file_format="xls", exists=exists)

    decision = choose_backend("create_new", inventory)

    assert decision.backend != BACKEND_XLSXWRITER
    assert decision.backend == BACKEND_EXCEL_REQUIRED
    assert decision.explain.get("file_format") == "xls"


def test_plain_xlsx_edit_routes_to_openpyxl() -> None:
    inventory = _inventory()

    decision = choose_backend("edit_existing", inventory)

    assert decision.backend == BACKEND_OPENPYXL


def test_plain_xlsm_with_no_risk_features_also_routes_to_openpyxl() -> None:
    inventory = _inventory(file_format="xlsm")

    decision = choose_backend("edit_existing", inventory)

    assert decision.backend == BACKEND_OPENPYXL


def test_macro_enabled_workbook_routes_to_excel_required_with_explain() -> None:
    inventory = _inventory(file_format="xlsm", has_macros=True)

    decision = choose_backend("edit_existing", inventory)

    assert decision.backend == BACKEND_EXCEL_REQUIRED
    assert decision.explain.get("has_macros") is True


def test_signed_workbook_routes_to_excel_required_with_explain() -> None:
    inventory = _inventory(is_signed=True)

    decision = choose_backend("edit_existing", inventory)

    assert decision.backend == BACKEND_EXCEL_REQUIRED
    assert decision.explain.get("is_signed") is True


def test_external_link_workbook_routes_to_excel_required_with_explain() -> None:
    inventory = _inventory(has_external_links=True)

    decision = choose_backend("edit_existing", inventory)

    assert decision.backend == BACKEND_EXCEL_REQUIRED
    assert decision.explain.get("has_external_links") is True


@pytest.mark.parametrize(
    "field", ["has_data_model", "has_pivots", "has_slicers", "has_embedded_objects"]
)
def test_each_remaining_risk_feature_alone_routes_to_excel_required(field) -> None:
    inventory = _inventory(**{field: True})

    decision = choose_backend("edit_existing", inventory)

    assert decision.backend == BACKEND_EXCEL_REQUIRED
    assert decision.explain.get(field) is True


def test_workbook_connection_routes_to_excel_required_with_explain() -> None:
    # Issue #70: a detected workbook connection (Power Query or legacy
    # QueryTable/OLEDB/ODBC) must fail closed exactly like the other seven
    # risk fields, regardless of every other field being unset.
    inventory = _inventory(has_connections=True)

    decision = choose_backend("edit_existing", inventory)

    assert decision.backend == BACKEND_EXCEL_REQUIRED
    assert decision.explain.get("has_connections") is True


def test_multiple_risk_features_are_all_named_in_explain() -> None:
    inventory = _inventory(has_macros=True, has_pivots=True)

    decision = choose_backend("edit_existing", inventory)

    assert decision.backend == BACKEND_EXCEL_REQUIRED
    assert decision.explain.get("has_macros") is True
    assert decision.explain.get("has_pivots") is True


def test_xls_edit_routes_to_convert_required() -> None:
    inventory = _inventory(file_format="xls")

    decision = choose_backend("edit_existing", inventory)

    assert decision.backend == BACKEND_CONVERT_REQUIRED


def test_csv_edit_is_declared_out_of_scope_not_excel_required() -> None:
    inventory = _inventory(file_format="csv")

    decision = choose_backend("edit_existing", inventory)

    assert decision.backend == BACKEND_NOT_APPLICABLE
    assert decision.backend != BACKEND_EXCEL_REQUIRED


def test_tsv_edit_is_declared_out_of_scope_not_excel_required() -> None:
    inventory = _inventory(file_format="tsv")

    decision = choose_backend("edit_existing", inventory)

    assert decision.backend == BACKEND_NOT_APPLICABLE


def test_edit_existing_nonexistent_file_fails_closed() -> None:
    inventory = _inventory(exists=False)

    decision = choose_backend("edit_existing", inventory)

    assert decision.backend == BACKEND_EXCEL_REQUIRED


def test_unclassifiable_workbook_fails_closed_regardless_of_intent() -> None:
    inventory = _inventory(is_classifiable=False)

    decision = choose_backend("edit_existing", inventory)

    assert decision.backend == BACKEND_EXCEL_REQUIRED
    assert decision.explain.get("is_classifiable") is False


def test_unclassifiable_workbook_fails_closed_even_with_no_other_risk_flags_set() -> None:
    # is_classifiable=False must win even though every risk boolean below it
    # defaults to False -- "not detected" must not be read as "confirmed safe".
    inventory = _inventory(is_classifiable=False, has_macros=False, has_pivots=False)

    decision = choose_backend("edit_existing", inventory)

    assert decision.backend == BACKEND_EXCEL_REQUIRED


def test_unmatched_intent_fails_closed_rather_than_defaulting_to_a_library() -> None:
    inventory = _inventory()

    decision = choose_backend("convert_format", inventory)

    assert decision.backend == BACKEND_EXCEL_REQUIRED


def test_unknown_intent_raises_value_error() -> None:
    inventory = _inventory()

    with pytest.raises(ValueError):
        choose_backend("delete_everything", inventory)


def test_decision_is_deterministic_for_the_same_inputs() -> None:
    inventory = _inventory(has_macros=True)

    first = choose_backend("edit_existing", inventory)
    second = choose_backend("edit_existing", inventory)

    assert first == second


def test_xlsxwriter_new_workbook_creation_is_faster_than_openpyxl(tmp_path) -> None:
    # Lightweight sanity timing comparison only, per the descope: not a
    # benchmark corpus (that belongs to #39). New-workbook creation with no
    # existing file to preserve is exactly the router's xlsxwriter fast
    # path above, so this is the one comparison the descope calls for.
    rows, cols = 3000, 12

    xlsxwriter_path = tmp_path / "xlsxwriter_out.xlsx"
    start = time.perf_counter()
    workbook = xlsxwriter.Workbook(str(xlsxwriter_path))
    sheet = workbook.add_worksheet()
    for row in range(rows):
        for col in range(cols):
            sheet.write_number(row, col, row * cols + col)
    workbook.close()
    xlsxwriter_seconds = time.perf_counter() - start

    openpyxl_path = tmp_path / "openpyxl_out.xlsx"
    start = time.perf_counter()
    wb = openpyxl.Workbook()
    ws = wb.active
    for row in range(1, rows + 1):
        for col in range(1, cols + 1):
            ws.cell(row=row, column=col, value=(row - 1) * cols + (col - 1))
    wb.save(openpyxl_path)
    openpyxl_seconds = time.perf_counter() - start

    assert xlsxwriter_path.exists()
    assert openpyxl_path.exists()
    assert xlsxwriter_seconds < openpyxl_seconds
