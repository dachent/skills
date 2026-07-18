"""Certification corpus generator for issue #39 (single-machine subset).

Generates a small, representative set of throwaway workbooks *on demand*
into a caller-supplied directory -- nothing here is a committed binary
fixture, matching the existing `tests/wb_fixtures.py` (Python) /
`FixtureWorkbookBuilder.cs` (C#) pattern elsewhere in this repo. Every
corpus item carries its own documented expected outcome (router decision,
whether it should be pushed through the real supervisor, and whether its
validation contract should pass or fail) so `run_corpus.py` has something
concrete to assert against, not just "did it not crash."

Five of the six items (everything except `power_query_minimal`) never
launch Excel and never shell out to anything -- built purely with openpyxl
and the stdlib `zipfile` module, including, for two items, injecting a
single placeholder/inert OOXML zip entry that
`control_plane/workbook_inventory.py`'s namelist-based detection keys on.
Neither of those two placeholder items needs its injected part to be a
genuinely valid, Excel-openable macro project or external-link relationship:
`#35`'s router only inspects the raw zip namelist (see
workbook_inventory.py's module docstring), so a placeholder entry at the
exact expected path is sufficient to exercise the *routing* decision that
item is meant to prove. Neither is ever opened by real Excel in this issue's
harness -- exactly because their injected parts are not real, so there is no
reason to risk finding out how Excel's repair/recovery UI reacts to them on
a machine the owner actively uses.

`power_query_minimal` is the one exception: a genuine Power Query M
connection's OOXML representation (`xl/connections.xml` plus a `customXml`
query-definition part) is not something this issue attempts to hand-craft
the way the macro/external-link placeholders above do -- instead,
`build_power_query_item` shells out to the repo's existing, already-tested
`xlsx-win/scripts/power_query_excel.ps1` (`upsert-query` then
`load-worksheet`) against a blank workbook, which means it genuinely
launches Excel (twice) to build this one fixture. Callers MUST have already
satisfied `excel_safety`'s preflight gate before calling it -- see
run_corpus.py's main(), which is also the only place that decides whether
this item is built at all.
"""

from __future__ import annotations

import subprocess
import zipfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

_LEGACY_POWER_QUERY_SCRIPT = (
    Path(__file__).resolve().parent.parent.parent / "scripts" / "power_query_excel.ps1"
)

# Marker text written into a dedicated, otherwise-unused cell in a couple of
# corpus items, purely so a validation contract has an unambiguous literal
# value to assert on (openpyxl never calculates formulas, so cached formula
# results are not a safe thing for a contract to assert against a
# corpus item that was only ever written by openpyxl, never opened by Excel).
_SENTINEL_MARKER = "generated-by-certification-corpus"


@dataclass(frozen=True)
class CorpusItem:
    """One corpus workbook plus its documented expected outcome.

    `exercise_supervisor` is an explicit, independent flag -- not derived
    from `expected_backend == "excel_required"`. See run_corpus.py and
    README.md for why: two of the `excel_required` items here are
    deliberately *not* pushed through the real supervisor (a macro-policy
    rejection and a router-only external-link check don't need Excel to
    prove their documented outcome, and one of them carries a placeholder
    part not safe to hand to real Excel -- see the module docstring above).
    """

    name: str
    input_path: Path
    intent: str
    expected_backend: str
    description: str
    exercise_supervisor: bool = False
    steps: tuple = ()
    contract: dict | None = None
    expected_contract_pass: bool | None = None
    expected_supervisor_ok: bool = True
    """Whether the supervisor job is expected to report ok=True. Default True
    (the ordinary case). power_query_minimal sets this False: on this
    machine, EXCEL.EXE reproducibly does not exit within any tested budget
    (60s, then 300s) after a genuine Power Query refresh, even though the
    actual work (refresh/calc/save) completes in seconds and produces a
    correct output file -- see its own description and README.md. The
    supervisor's TIMED_OUT verdict in that case is the CORRECT, conservative
    behavior (it cannot confirm clean process teardown, so it does not claim
    success), not a bug -- but a caller of this harness needs a way to say
    "TIMED_OUT is the expected, documented outcome here," which is what this
    field is for."""
    macro_check: tuple | None = None  # (macro_name, allowlist) for control_plane.macro_policy


def _build_plain_workbook(path: Path) -> None:
    """A plain xlsx: formulas, no macros/signature/links/data model/pivots/
    slicers/embedded objects. Expected router decision: openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["ID", "Quantity", "Price"])
    ws.append([1, 10, 2.5])
    ws.append([2, 5, 4.0])
    ws.append([3, 8, 1.25])
    ws["E2"] = "=B2*C2"
    ws["E3"] = "=B3*C3"
    ws["E4"] = "=B4*C4"
    ws["G1"] = _SENTINEL_MARKER
    wb.save(path)


def _build_macro_workbook(path: Path) -> None:
    """An .xlsm with a placeholder xl/vbaProject.bin zip entry.

    Not a real compiled VBA project -- fabricating genuinely compiled VBA
    bytecode from scratch, without Excel or a third-party OLE-writer
    library, is out of scope for this issue. workbook_inventory.py's
    has_macros detection is an exact zip-namelist match on
    "xl/vbaProject.bin" (see its module docstring: it inspects the raw
    package, never opens it with openpyxl or Excel), so a placeholder blob
    at that exact path is sufficient to trigger the router decision this
    item exists to prove. This file's content types are not adjusted to
    declare a macro-enabled workbook, and it is never opened by Excel COM
    in this issue's harness -- see run_corpus.py, which checks macro
    rejection at the control_plane.macro_policy layer instead of by
    actually invoking the supervisor's (unimplemented, per #36's own
    README) run_approved_macro step.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Macro-enabled corpus fixture (placeholder VBA project; not executable)."
    wb.save(path)

    with zipfile.ZipFile(path, "a", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("xl/vbaProject.bin", b"placeholder-not-a-real-compiled-vba-project")


def _build_table_workbook(path: Path) -> None:
    """A plain xlsx with one worksheet Table (ListObject) holding two data
    rows. No workbook connection -- a real live Power-Query-backed
    connection refresh is already proven by the C# integration test
    PerConnectionRefreshTests (#36); this item's purpose (see README.md) is
    narrower: prove run_corpus.py's own pipeline correctly drives a
    `refresh` step end-to-end against a Table-bearing workbook through the
    real supervisor, not to duplicate live-connection-content fidelity
    proof. Expected router decision: openpyxl (no risk feature this
    router tracks is present -- workbook connections are not one of the
    seven tracked risk fields; see README.md's "known gap" note)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["Region", "Amount"])
    ws.append(["East", 100])
    ws.append(["West", 200])

    table = Table(displayName="SalesData", ref="A1:B3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

    wb.save(path)


def _build_external_link_workbook(path: Path) -> None:
    """A plain xlsx plus one inert, unreferenced xl/externalLinks/*.xml zip
    entry. Expected router decision: excel_required.

    The injected part is not wired into any real relationship/content-type
    declaration -- it exists purely so workbook_inventory.py's prefix match
    on "xl/externallinks/" fires, exactly like the macro item above.
    Deliberately never exercised through the real supervisor in
    run_corpus.py (this item's only documented expected outcome is the
    router decision) -- there is no reason to find out how real Excel's
    repair/recovery UI would react to a placeholder external-link part on a
    machine the owner actively uses, and the issue's own corpus item
    description for this entry makes no claim beyond the router decision.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Workbook with an inert, placeholder external-workbook-link OOXML part."
    wb.save(path)

    with zipfile.ZipFile(path, "a", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(
            "xl/externalLinks/externalLink1.xml",
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            b'<externalLink xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"/>',
        )


def _build_failing_contract_workbook(path: Path) -> None:
    """A plain, otherwise-unremarkable xlsx paired (by run_corpus.py) with a
    validation contract designed to fail -- proving the certification
    harness actually detects and reports a failure, not just happy paths."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["ID", "Value"])
    ws.append([1, 100])
    wb.save(path)


def _run_legacy_power_query_script(
    workbook_path: Path,
    action: str,
    query_name: str,
    *,
    mformula_path: Path | None = None,
    worksheet_name: str | None = None,
) -> None:
    """Invoke the existing xlsx-win/scripts/power_query_excel.ps1 as a
    subprocess. Launches Excel. Caller must have already satisfied
    excel_safety's preflight gate."""
    args = [
        "powershell",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(_LEGACY_POWER_QUERY_SCRIPT),
        "-WorkbookPath",
        str(workbook_path),
        "-Action",
        action,
        "-QueryName",
        query_name,
    ]
    if mformula_path is not None:
        args += ["-MFormulaPath", str(mformula_path)]
    if worksheet_name is not None:
        args += ["-WorksheetName", worksheet_name]

    result = subprocess.run(args, capture_output=True, text=True, timeout=180, check=False)
    if result.returncode != 0:
        raise RuntimeError(
            f"power_query_excel.ps1 -Action {action} failed (exit {result.returncode}): "
            f"stdout={result.stdout!r} stderr={result.stderr!r}"
        )


def build_power_query_item(directory: str | Path) -> CorpusItem:
    """Build the one corpus item with a genuine Power Query M connection, by
    shelling out to the existing power_query_excel.ps1 (see module
    docstring). Launches Excel. Caller MUST have already satisfied
    excel_safety's preflight gate -- this function performs no safety check
    of its own and is not called by build_corpus().

    Expected router decision: openpyxl. This is the SAME known gap as
    table_connection above (workbook connections are not one of the seven
    risk fields workbook_inventory.py tracks) -- but a more consequential
    instance of it: SKILL.md's own existing guidance is explicit that Power
    Query M work must go through Excel COM, never file-only libraries, so
    routing a genuinely M-code-backed workbook to openpyxl for an "edit" is
    not merely a suboptimal choice but contrary to the skill's own
    documented rule. Confirmed by direct measurement against this exact
    workbook, not assumed. Tracked as a real, open gap against #35 in
    README.md -- fixing the router is separate scope from this issue.
    """
    directory = Path(directory)
    directory.mkdir(parents=True, exist_ok=True)

    path = directory / "power_query_minimal.xlsx"
    # power_query_excel.ps1 opens an existing workbook (Workbooks.Open) --
    # it does not create one -- so a blank workbook must exist first.
    Workbook().save(path)

    m_formula = (
        "let\n"
        "    Source = Table.FromRows(\n"
        '        {{1, "Alpha"}, {2, "Beta"}, {3, "Gamma"}},\n'
        '        {"Id", "Name"}\n'
        "    )\n"
        "in\n"
        "    Source"
    )
    m_path = directory / "power_query_minimal.m"
    m_path.write_text(m_formula, encoding="utf-8")

    _run_legacy_power_query_script(path, "upsert-query", "SyntheticSource", mformula_path=m_path)
    _run_legacy_power_query_script(path, "load-worksheet", "SyntheticSource", worksheet_name="Loaded")

    return CorpusItem(
        name="power_query_minimal",
        input_path=path,
        intent="edit_existing",
        expected_backend="openpyxl",
        description=(
            "Workbook with a genuine, self-contained Power Query M connection "
            "(Table.FromRows -- no external file or network dependency), built via the "
            "existing power_query_excel.ps1 rather than a hand-crafted placeholder part. "
            "Expected (current, undesirable) router decision: openpyxl -- the same known "
            "gap as table_connection, now confirmed against a genuine M-code-backed "
            "connection rather than only a plain Table; see README.md's 'known gap' "
            "section for why this is a more consequential instance of it.\n\n"
            "MAJOR FINDING, then FIXED and reverified: the first two real runs against this "
            "item (60s and 300s refresh/close budgets) reproducibly hit a genuine defect -- "
            "the job's actual work (open, refresh, recalc, save) completed in ~9 seconds "
            "every time (confirmed via events.jsonl reaching SAVING then SUCCEEDED, and a "
            "correct output.xlsx), but EXCEL.EXE itself never exited within either budget, "
            "so the supervisor correctly timed out and force-killed via the Job Object "
            "(zero orphaned processes both times). Root cause, found via web research and "
            "confirmed by code inspection, not guessed: StepRunner.cs's RunRefresh obtained "
            "Workbook.Connections, each Connection, and each OLEDBConnection/ODBCConnection "
            "via COM property access and never released any of them -- classic unreleased-"
            "RCW behavior (Application.Quit() only requests an exit; Excel won't actually "
            "terminate until every outstanding COM reference is released, and relying on the "
            ".NET GC alone to get there can take multiple collection cycles, or apparently "
            "never resolve within any tested budget for a genuine Power-Query/Mashup-backed "
            "connection specifically). Fixed by explicit Marshal.ReleaseComObject on every "
            "one of those intermediate objects (in finally blocks, surviving exceptions) "
            "plus strengthening ExcelSession.cs's single GC.Collect()+WaitForPendingFinalizers "
            "to the standard double-collect pattern. Reverified against this exact item after "
            "the fix: SUCCEEDED, ok=True, in 12.5 seconds -- down from timing out past 300s. "
            "See README.md for the full writeup and sources."
        ),
        exercise_supervisor=True,
        steps=(
            {"type": "open", "workbook_path": str(path)},
            {"type": "refresh", "connections": "all"},
            {"type": "recalc", "mode": "full_rebuild"},
        ),
        contract={
            "required_sheets": ["Loaded"],
            "min_row_counts": {"Loaded": 4},
            "sentinel_cells": [{"sheet": "Loaded", "cell": "B2", "expected": "Alpha"}],
        },
        expected_contract_pass=True,
    )


def build_corpus(directory: str | Path) -> list[CorpusItem]:
    """Generate every corpus workbook into `directory` (created if needed)
    and return the list of CorpusItem descriptors, each carrying its own
    documented expected outcome. Never touches Excel."""
    directory = Path(directory)
    directory.mkdir(parents=True, exist_ok=True)

    plain_path = directory / "plain_formulas.xlsx"
    _build_plain_workbook(plain_path)

    macro_path = directory / "macro_enabled.xlsm"
    _build_macro_workbook(macro_path)

    table_path = directory / "table_connection.xlsx"
    _build_table_workbook(table_path)

    external_link_path = directory / "external_link.xlsx"
    _build_external_link_workbook(external_link_path)

    failing_contract_path = directory / "failing_contract.xlsx"
    _build_failing_contract_workbook(failing_contract_path)

    return [
        CorpusItem(
            name="plain_formulas",
            input_path=plain_path,
            intent="edit_existing",
            expected_backend="openpyxl",
            description=(
                "Plain workbook, formulas, no risk features. Expected router decision: "
                "openpyxl. Should pass a simple validation contract evaluated directly "
                "against this file (no Excel involved)."
            ),
            exercise_supervisor=False,
            contract={
                "required_sheets": ["Data"],
                "min_row_counts": {"Data": 4},
                "sentinel_cells": [{"sheet": "Data", "cell": "G1", "expected": _SENTINEL_MARKER}],
            },
            expected_contract_pass=True,
        ),
        CorpusItem(
            name="macro_enabled",
            input_path=macro_path,
            intent="edit_existing",
            expected_backend="excel_required",
            description=(
                "Macro-enabled workbook (.xlsm with a placeholder xl/vbaProject.bin part). "
                "Expected router decision: excel_required. This item's macro_check only "
                "proves two narrower things, and does NOT prove Excel-level macro "
                "rejection or that a genuinely macro-bearing workbook is treated "
                "differently from a plain one at execution time: (1) "
                "workbook_inventory.inspect_workbook(...).has_macros is True for this "
                "file (the router's zip-namelist detection actually fired), and (2) "
                "macro_policy.is_macro_approved(...) -- a pure sha256+entrypoint lookup "
                "against a caller-supplied allowlist, with no awareness of the "
                "workbook's actual content -- returns False against an empty allowlist, "
                "which it would do identically for any input file, macro-bearing or not. "
                "The supervisor's own run_approved_macro step (the one place that would "
                "actually gate execution at the Excel/COM level) is unimplemented -- see "
                "supervisor/README.md ('MACRO_EXECUTION_DEFERRED') -- and this corpus item "
                "does not exercise it. 'Prove Excel-level AutomationSecurity macro "
                "rejection' remains an open gap, tracked here and in README.md, not "
                "something this check stands in for."
            ),
            exercise_supervisor=False,
            macro_check=("TestMacro", []),
        ),
        CorpusItem(
            name="table_connection",
            input_path=table_path,
            intent="edit_existing",
            expected_backend="openpyxl",
            description=(
                "Worksheet Table with two data rows, no live connection. Expected router "
                "decision: openpyxl (workbook connections are not a tracked risk field -- "
                "see README.md's 'known gap' note). Deliberately still exercised through "
                "the real supervisor (open, refresh(all), recalc, save_as) regardless of "
                "the router's decision, because this item's documented purpose is to "
                "prove the pipeline drives a refresh step end-to-end against a "
                "Table-bearing workbook -- not to test routing. The output is then "
                "validated against a contract requiring the Table and its row count."
            ),
            exercise_supervisor=True,
            steps=(
                {"type": "open", "workbook_path": str(table_path)},
                {"type": "refresh", "connections": "all"},
                {"type": "recalc", "mode": "full_rebuild"},
            ),
            contract={
                "required_tables": ["SalesData"],
                "min_row_counts": {"SalesData": 2},
                "sentinel_cells": [{"sheet": "Sales", "cell": "B2", "expected": 100}],
            },
            expected_contract_pass=True,
        ),
        CorpusItem(
            name="external_link",
            input_path=external_link_path,
            intent="edit_existing",
            expected_backend="excel_required",
            description=(
                "Workbook with an external workbook link (placeholder OOXML part). "
                "Expected router decision: excel_required. Router-decision-only -- never "
                "pushed through real Excel in this harness (see this module's build "
                "function for why)."
            ),
            exercise_supervisor=False,
        ),
        CorpusItem(
            name="failing_contract",
            input_path=failing_contract_path,
            intent="edit_existing",
            expected_backend="openpyxl",
            description=(
                "Otherwise-plain workbook paired with a validation contract designed to "
                "fail (a row-count minimum far above the actual data), to prove the "
                "certification harness actually detects and reports a failure rather "
                "than only exercising happy paths."
            ),
            exercise_supervisor=False,
            contract={
                "required_sheets": ["Data"],
                "min_row_counts": {"Data": 50},
            },
            expected_contract_pass=False,
        ),
    ]
