"""Independent fresh-package proof for a completed composite transaction."""

from __future__ import annotations

import hashlib
import json
import zipfile
from decimal import Decimal
from pathlib import Path

from jsonschema.validators import validator_for
from lxml import etree
from openpyxl.formula.translate import Translator

from .errors import ContractError
from .ooxml_table_transaction import (
    MAIN,
    NS,
    PackagePreflight,
    _column_letters,
    _find_sheet_part,
    _find_table_part,
    _linked_pivots,
    _parse_cell,
    _parse_range,
    _saved_sort,
)
from .schemas import SCHEMA_DIR
from .table_sidecar import iter_typed_rows

_ORACLE_SCHEMA = json.loads((SCHEMA_DIR / "pivot-oracle-v1.schema.json").read_text(encoding="utf-8"))
_ORACLE_VALIDATOR = validator_for(_ORACLE_SCHEMA)(_ORACLE_SCHEMA)


def sha256_path(path: str | Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _shared_strings(package: zipfile.ZipFile) -> list[str]:
    try:
        root = etree.fromstring(package.read("xl/sharedStrings.xml"))
    except KeyError:
        return []
    return ["".join(node.itertext()) for node in root.findall(f"{{{MAIN}}}si")]


def _normalize_number(value: str) -> str:
    normalized = format(Decimal(value).normalize(), "f")
    return "0" if normalized == "-0" else normalized


def _numbers_equivalent(expected: str | None, actual: str | None) -> bool:
    if expected is None or actual is None:
        return expected is actual
    try:
        expected_decimal = Decimal(expected)
        actual_decimal = Decimal(actual)
        quantum = Decimal(1).scaleb(expected_decimal.as_tuple().exponent)
        return actual_decimal.quantize(quantum) == expected_decimal
    except Exception:
        return False

def _translate_formula(formula: str, origin: str, destination: str) -> str:
    try:
        translated = Translator(f"={formula}", origin=origin).translate_formula(destination)
    except Exception as exc:
        raise ContractError(
            "RESULT_PROOF_INVALID",
            "A calculated-column formula could not be normalized.",
            {"formula": formula, "origin": origin, "destination": destination},
        ) from exc
    return translated[1:] if translated.startswith("=") else translated


def _cell_record(
    cell: etree._Element,
    shared: list[str],
    shared_formulas: dict[str, tuple[str, str]],
    reference: str,
) -> dict:
    cell_type = cell.get("t")
    formula_node = cell.find(f"{{{MAIN}}}f")
    formula = formula_node.text if formula_node is not None else None
    if formula_node is not None and formula_node.get("t") == "shared":
        shared_id = formula_node.get("si")
        if formula and shared_id is not None:
            shared_formulas[shared_id] = (formula, reference)
        elif shared_id in shared_formulas:
            formula, origin = shared_formulas[shared_id]
            formula = _translate_formula(formula, origin, reference)
        else:
            raise ContractError("RESULT_PROOF_INVALID", "Shared formula base is unavailable.", {"cell": reference})
    if cell_type == "inlineStr":
        value = "".join(cell.find(f"{{{MAIN}}}is").itertext()) if cell.find(f"{{{MAIN}}}is") is not None else ""
        storage = "inline_string"
    elif cell_type == "s":
        raw = cell.findtext(f"{{{MAIN}}}v") or "0"
        value = shared[int(raw)]
        storage = "inline_string"
    elif cell_type == "b":
        value = "true" if cell.findtext(f"{{{MAIN}}}v") == "1" else "false"
        storage = "boolean"
    elif cell_type == "str":
        # Formula cells with cached text results use the OOXML `str` type.
        value = cell.findtext(f"{{{MAIN}}}v") or ""
        storage = "inline_string"
    else:
        raw = cell.findtext(f"{{{MAIN}}}v")
        value = _normalize_number(raw) if raw not in {None, ""} else None
        storage = "number" if value is not None else "blank"
    return {"storage": storage, "value": value, "formula": formula, "style": cell.get("s")}


def _iter_rows(
    package: zipfile.ZipFile,
    worksheet_part: str,
    first_row: int,
    last_row: int,
    first_column: int,
    last_column: int,
):
    shared = _shared_strings(package)
    shared_formulas: dict[str, tuple[str, str]] = {}
    with package.open(worksheet_part, "r") as stream:
        for _event, element in etree.iterparse(stream, events=("end",), tag=f"{{{MAIN}}}row", huge_tree=True):
            row_number = int(element.get("r", "0"))
            if first_row <= row_number <= last_row:
                values = {}
                for cell in element.findall(f"{{{MAIN}}}c"):
                    reference = cell.get("r")
                    if not reference:
                        continue
                    column, _ = _parse_cell(reference)
                    if first_column <= column <= last_column:
                        values[column] = _cell_record(cell, shared, shared_formulas, reference)
                yield row_number, values
            element.clear()
            while element.getprevious() is not None:
                del element.getparent()[0]
            if row_number > last_row:
                break


class _RowCursor:
    def __init__(self, iterator):
        self._iterator = iter(iterator)
        self._current = next(self._iterator, None)

    def row(self, number: int) -> dict:
        while self._current is not None and self._current[0] < number:
            self._current = next(self._iterator, None)
        if self._current is not None and self._current[0] == number:
            result = self._current[1]
            self._current = next(self._iterator, None)
            return result
        return {}


def _expected_record(cell) -> tuple[str, str | None]:
    if cell.normalized is None:
        return "blank", None
    return cell.storage_type, cell.normalized


def _verify_body(
    seed: zipfile.ZipFile,
    output: zipfile.ZipFile,
    operation: dict,
    schema: dict,
    original: PackagePreflight,
    output_sheet_part: str,
) -> dict:
    body_start = original.body_start_row
    old_end = original.old_body_end_row
    final_end = original.header_row + operation["table"]["final_body_rows"]
    scan_end = max(old_end, final_end)
    output_rows = _RowCursor(_iter_rows(output, output_sheet_part, body_start, scan_end, original.first_column, original.last_column))
    seed_rows = _RowCursor(_iter_rows(seed, original.worksheet_part, body_start, old_end, original.first_column, original.last_column))
    source_rows = iter(iter_typed_rows(operation["source"]["path"], schema))
    source_start = old_end + 1 if operation["type"] == "append_table_rows" else body_start
    source_end = final_end
    formula_by_column: dict[int, tuple[str, str]] = {}
    prefix_digest = hashlib.sha256()
    source_digest = hashlib.sha256()
    formula_cells = 0

    for row_number in range(body_start, scan_end + 1):
        actual = output_rows.row(row_number)
        if operation["type"] == "append_table_rows" and row_number <= old_end:
            before = seed_rows.row(row_number)
            for offset, column in enumerate(operation["table"]["columns"]):
                if column["role"] != "writable":
                    continue
                column_number = original.first_column + offset
                left = before.get(column_number, {"storage": "blank", "value": None, "style": None})
                right = actual.get(column_number, {"storage": "blank", "value": None, "style": None})
                comparison = [left["storage"], left["value"], left["style"]]
                if comparison != [right["storage"], right["value"], right["style"]]:
                    raise ContractError("RESULT_PROOF_INVALID", "Existing append prefix changed.", {"row": row_number, "column": column["name"]})
                prefix_digest.update(json.dumps([row_number, column_number, comparison], separators=(",", ":")).encode())

        if source_start <= row_number <= source_end:
            try:
                expected_row = next(source_rows)
            except StopIteration as exc:
                raise ContractError("RESULT_PROOF_INVALID", "Source ended before final Table geometry.", {}) from exc
            for offset, expected_cell in enumerate(expected_row):
                if expected_cell.role != "writable":
                    continue
                column_number = original.first_column + offset
                found = actual.get(column_number, {"storage": "blank", "value": None})
                expected_storage, expected_value = _expected_record(expected_cell)
                values_match = (found["storage"], found["value"]) == (expected_storage, expected_value)
                if not values_match and found["storage"] == expected_storage == "number":
                    values_match = _numbers_equivalent(expected_value, found["value"])
                if not values_match:
                    raise ContractError(
                        "RESULT_PROOF_INVALID", "Saved typed value differs from the immutable source.",
                        {"row": row_number, "column": expected_cell.name, "expected": [expected_storage, expected_value], "actual": [found["storage"], found["value"]]},
                    )
                source_digest.update(json.dumps([row_number, column_number, expected_storage, expected_value], separators=(",", ":"), ensure_ascii=False).encode("utf-8"))

        if body_start <= row_number <= final_end:
            for offset, column in enumerate(operation["table"]["columns"]):
                if column["role"] != "calculated":
                    continue
                column_number = original.first_column + offset
                formula = actual.get(column_number, {}).get("formula")
                if not formula:
                    raise ContractError("RESULT_PROOF_INVALID", "Calculated-column formula is missing.", {"row": row_number, "column": column["name"]})
                reference = f"{_column_letters(column_number)}{row_number}"
                if column_number not in formula_by_column:
                    formula_by_column[column_number] = (formula, reference)
                else:
                    prior, origin = formula_by_column[column_number]
                    if formula != _translate_formula(prior, origin, reference):
                        raise ContractError("RESULT_PROOF_INVALID", "Calculated column contains a formula exception.", {"row": row_number, "column": column["name"]})
                formula_cells += 1

        if operation["type"] == "replace_table_data" and final_end < row_number <= old_end and actual:
            raise ContractError("RESULT_PROOF_INVALID", "Stale shrink-tail cells remain outside the final Table.", {"row": row_number, "cells": sorted(actual)})

    try:
        next(source_rows)
        raise ContractError("RESULT_PROOF_INVALID", "Source has extra rows beyond final Table geometry.", {})
    except StopIteration:
        pass
    return {
        "existing_prefix_sha256": prefix_digest.hexdigest() if operation["type"] == "append_table_rows" else None,
        "source_rows_sha256": source_digest.hexdigest(),
        "formula_columns": len(formula_by_column),
        "formula_cells": formula_cells,
        "old_tail_absent": True,
    }


def _matrix_hash(package: zipfile.ZipFile, sheet_part: str, reference: str) -> tuple[str, int, int]:
    first_col, first_row, last_col, last_row = _parse_range(reference)
    cursor = _RowCursor(_iter_rows(package, sheet_part, first_row, last_row, first_col, last_col))
    digest = hashlib.sha256()
    for row in range(first_row, last_row + 1):
        values = cursor.row(row)
        for column in range(first_col, last_col + 1):
            cell = values.get(column, {"storage": "blank", "value": None, "formula": None})
            digest.update(json.dumps([row, column, cell["storage"], cell["value"], cell["formula"]], separators=(",", ":"), ensure_ascii=False).encode("utf-8"))
    return digest.hexdigest(), last_row - first_row + 1, last_col - first_col + 1


def _verify_pivot_structure(package: zipfile.ZipFile, workbook: etree._Element, operation: dict) -> dict:
    caches, reports = _linked_pivots(package, workbook, operation["table"]["name"])
    expected_records = operation["table"]["final_body_rows"]
    cache_results = []
    for part in caches:
        root = etree.fromstring(package.read(part))
        try:
            record_count = int(root.get("recordCount", ""))
        except ValueError as exc:
            raise ContractError("RESULT_PROOF_INVALID", "Pivot cache recordCount is missing or invalid.", {"part": part}) from exc
        if record_count != expected_records:
            raise ContractError(
                "RESULT_PROOF_INVALID",
                "Pivot cache recordCount differs from the final Table body row count.",
                {"part": part, "actual": record_count, "expected": expected_records},
            )
        cache_results.append({"part": part, "record_count": record_count})

    report_results = []
    for part in reports:
        root = etree.fromstring(package.read(part))
        location = root.find("m:location", NS)
        reference = location.get("ref") if location is not None else None
        if not reference:
            raise ContractError("RESULT_PROOF_INVALID", "Pivot report has no saved location.", {"part": part})
        report_results.append({"part": part, "name": root.get("name"), "location": reference})
    return {"mode": "topology_only", "caches": cache_results, "reports": report_results}


def _verify_oracle(package: zipfile.ZipFile, workbook: etree._Element, operation: dict) -> dict:
    pivot_spec = operation["dependent_pivots"]
    oracle_path_value = pivot_spec.get("oracle_path")
    oracle_sha256 = pivot_spec.get("oracle_sha256")
    if bool(oracle_path_value) != bool(oracle_sha256):
        raise ContractError("RESULT_PROOF_INVALID", "Pivot oracle path/hash must be supplied together.", {})
    if not oracle_path_value:
        return _verify_pivot_structure(package, workbook, operation)
    oracle_path = Path(oracle_path_value)
    if sha256_path(oracle_path).lower() != oracle_sha256.lower():
        raise ContractError("RESULT_PROOF_INVALID", "Pivot oracle hash changed.", {})
    oracle = json.loads(oracle_path.read_text(encoding="utf-8"))
    try:
        _ORACLE_VALIDATOR.validate(oracle)
    except Exception as exc:
        raise ContractError("RESULT_PROOF_INVALID", "Pivot oracle is invalid.", {"error": str(exc)}) from exc
    if len(oracle["reports"]) != operation["dependent_pivots"]["report_count"]:
        raise ContractError("RESULT_PROOF_INVALID", "Pivot oracle report count is incomplete.", {})
    results = []
    for report in oracle["reports"]:
        sheet_part = _find_sheet_part(package, workbook, report["sheet"])
        observed_hash, rows, columns = _matrix_hash(package, sheet_part, report["range"])
        if observed_hash.lower() != report["matrix_sha256"].lower() or rows != report["rows"] or columns != report["columns"]:
            raise ContractError("RESULT_PROOF_INVALID", "Pivot report matrix differs from its independent oracle.", {"report": report["name"], "observed_sha256": observed_hash})
        results.append({"name": report["name"], "sha256": observed_hash, "rows": rows, "columns": columns})
    return {"mode": "matrix_oracle", "reports": results}


def verify_final_package(
    seed_path: str | Path,
    output_path: str | Path,
    operation: dict,
    sidecar_schema: dict,
    original_preflight: PackagePreflight,
    expected_seed_sha256: str,
) -> dict:
    if sha256_path(seed_path).lower() != expected_seed_sha256.lower():
        raise ContractError("RESULT_PROOF_INVALID", "Seed workbook changed during the transaction.", {})
    with zipfile.ZipFile(seed_path, "r") as seed, zipfile.ZipFile(output_path, "r") as output:
        workbook = etree.fromstring(output.read("xl/workbook.xml"))
        output_sheet = _find_sheet_part(output, workbook, operation["table"]["sheet"])
        table_part, table = _find_table_part(output, output_sheet, operation["table"]["name"])
        expected_ref = (
            f"{_column_letters(original_preflight.first_column)}{original_preflight.header_row}:"
            f"{_column_letters(original_preflight.last_column)}{original_preflight.header_row + operation['table']['final_body_rows']}"
        )
        if table.get("ref") != expected_ref:
            raise ContractError("RESULT_PROOF_INVALID", "Final Table geometry is not exact.", {"actual": table.get("ref"), "expected": expected_ref})
        names = [node.get("name") for node in table.findall("m:tableColumns/m:tableColumn", NS)]
        if names != [column["name"] for column in operation["table"]["columns"]]:
            raise ContractError("RESULT_PROOF_INVALID", "Table identity/header schema changed.", {})
        observed_sort = _saved_sort(table, names)
        if observed_sort != operation["table"].get("saved_sort"):
            raise ContractError("RESULT_PROOF_INVALID", "Saved sort descriptor changed.", {})
        caches, reports = _linked_pivots(output, workbook, operation["table"]["name"])
        if (len(caches), len(reports)) != (
            operation["dependent_pivots"]["cache_count"], operation["dependent_pivots"]["report_count"]
        ):
            raise ContractError("RESULT_PROOF_INVALID", "Final linked Pivot topology changed.", {})
        body = _verify_body(seed, output, operation, sidecar_schema, original_preflight, output_sheet)
        oracle = _verify_oracle(output, workbook, operation)
        return {
            "schema_version": "1.0",
            "seed_sha256": expected_seed_sha256.lower(),
            "output_sha256": sha256_path(output_path),
            "table_part": table_part,
            "table_ref": table.get("ref"),
            "linked_cache_parts": list(caches),
            "linked_pivot_parts": list(reports),
            "body": body,
            "pivot_oracle": oracle,
        }
