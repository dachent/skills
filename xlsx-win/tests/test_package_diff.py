"""Package-diff test: an openpyxl edit must not silently drop OOXML parts.

Builds a small workbook (two sheets, a named range, a comment) with
openpyxl, copies it, makes a trivial edit via openpyxl on the copy, then
compares the zip's part namelist before and after. Some part-name/
timestamp/order churn from openpyxl re-serializing the whole package on
every save is expected and fine (docProps timestamps, calcChain, part
order) -- this test does not assert byte-identical zips. What it asserts is
that the *meaningful* parts -- each worksheet, the defined name, and the
comment -- survive the round trip. Verification reads the raw XML directly
with the standard library, the same way workbook_inventory.py does, rather
than asking openpyxl's own object model whether its own edit preserved
things.
"""

from __future__ import annotations

import shutil
import zipfile
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _worksheet_part_names(namelist: list) -> set:
    return {name for name in namelist if name.lower().startswith("xl/worksheets/sheet")}


def _has_comments_part(namelist: list) -> bool:
    return any(name.lower().startswith("xl/comments") for name in namelist)


def _defined_names(zf: zipfile.ZipFile) -> set:
    root = ET.fromstring(zf.read("xl/workbook.xml"))
    return {
        element.get("name")
        for element in root.iter()
        if _local_name(element.tag) == "definedName" and element.get("name")
    }


def _sheet_names(zf: zipfile.ZipFile) -> set:
    root = ET.fromstring(zf.read("xl/workbook.xml"))
    return {
        element.get("name")
        for element in root.iter()
        if _local_name(element.tag) == "sheet" and element.get("name")
    }


def test_openpyxl_edit_preserves_worksheets_defined_names_and_comments(tmp_path) -> None:
    original = tmp_path / "original.xlsx"

    wb = openpyxl.Workbook()
    data_sheet = wb.active
    data_sheet.title = "Data"
    data_sheet["A1"] = "value"
    data_sheet["A1"].comment = Comment("a note", "tester")
    wb.create_sheet("Summary")["A1"] = "=Data!A1"
    wb.defined_names["MyRange"] = DefinedName("MyRange", attr_text="Data!$A$1")
    wb.save(original)

    edited = tmp_path / "edited.xlsx"
    shutil.copy(original, edited)

    wb2 = openpyxl.load_workbook(edited)
    wb2["Data"]["B1"] = "a trivial edit"
    wb2.save(edited)

    with zipfile.ZipFile(original) as zf_before:
        namelist_before = zf_before.namelist()
        sheet_names_before = _sheet_names(zf_before)
        defined_names_before = _defined_names(zf_before)
        has_comments_before = _has_comments_part(namelist_before)
        worksheet_count_before = len(_worksheet_part_names(namelist_before))

    with zipfile.ZipFile(edited) as zf_after:
        namelist_after = zf_after.namelist()
        sheet_names_after = _sheet_names(zf_after)
        defined_names_after = _defined_names(zf_after)
        has_comments_after = _has_comments_part(namelist_after)
        worksheet_count_after = len(_worksheet_part_names(namelist_after))

    # Sanity: the fixture actually has the parts this test exists to guard.
    assert sheet_names_before == {"Data", "Summary"}
    assert defined_names_before == {"MyRange"}
    assert has_comments_before is True
    assert worksheet_count_before == 2

    # The trivial edit must not have dropped any of them.
    assert sheet_names_after == sheet_names_before
    assert defined_names_after == defined_names_before
    assert has_comments_after == has_comments_before
    assert worksheet_count_after == worksheet_count_before
