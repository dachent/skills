"""Certification corpus generator for issue #39 (single-machine subset).

Generates a small, representative set of throwaway workbooks *on demand*
into a caller-supplied directory -- nothing here is a committed binary
fixture, matching the existing `tests/wb_fixtures.py` (Python) /
`FixtureWorkbookBuilder.cs` (C#) pattern elsewhere in this repo. Every
corpus item carries its own documented expected outcome (router decision,
whether it should be pushed through the real supervisor, and whether its
validation contract should pass or fail) so `run_corpus.py` has something
concrete to assert against, not just "did it not crash."

Five of the six items (everything except `power_query_minimal`) never
launch Excel and never shell out to anything -- built purely with openpyxl
and the stdlib `zipfile` module, including, for three items, injecting a
single placeholder/inert OOXML zip entry that
`control_plane/workbook_inventory.py`'s namelist-based detection keys on.
None of those three placeholder items needs its injected part to be a
genuinely valid, Excel-openable macro project, external-link relationship,
or workbook connection: `#35`/`#70`'s router only inspects the raw zip
namelist (see workbook_inventory.py's module docstring), so a placeholder
entry at the exact expected path is sufficient to exercise the *routing*
decision that item is meant to prove. None of the three is ever opened by
real Excel in this issue's harness -- exactly because their injected parts
are not real, so there is no reason to risk finding out how Excel's
repair/recovery UI reacts to them on a machine the owner actively uses.

`power_query_minimal` is the one exception: a genuine Power Query M
connection's OOXML representation (`xl/connections.xml` plus a `customXml`
query-definition part) is not something this issue attempts to hand-craft
the way the macro/external-link/table-connection placeholders above do --
instead, `build_power_query_item` shells out to
`fixtures/build_power_query_fixture.ps1` (a small, test-fixture-only helper;
see issue #78 -- it is NOT part of the shipped skill, since Power Query M
authoring is a documented gap in the v2 job contract) against a blank
workbook, which means it genuinely launches Excel once to build this one
fixture. Callers MUST have already satisfied `excel_safety`'s preflight gate
before calling it -- see run_corpus.py's main(), which is also the only
place that decides whether this item is built at all.
"""

from __future__ import annotations

import subprocess
import zipfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

_POWER_QUERY_FIXTURE_SCRIPT = Path(__file__).resolve().parent / "fixtures" / "build_power_query_fixture.ps1"

# Marker text written into a dedicated, otherwise-unused cell in a couple of
# corpus items, purely so a validation contract has an unambiguous literal
# value to assert on (openpyxl never calculates formulas, so cached formula
# results are not a safe thing for a contract to assert against a
# corpus item that was only ever written by openpyxl, never opened by Excel).
_SENTINEL_MARKER = "generated-by-certification-corpus"


@dataclass(frozen=True)
class CorpusItem:
    """One corpus workbook plus its documented expected outcome.

    `exercise_supervisor` is an explicit, independent flag -- not derived
    from `expected_backend == "excel_required"`. See run_corpus.py and
    README.md for why: three of the `excel_required` items here
    (`macro_enabled`, `table_connection`, `external_link`) are deliberately
    *not* pushed through the real supervisor -- each carries a placeholder
    part not safe to hand to real Excel (see the module docstring above),
    and a macro-policy rejection / router-only connection or external-link
    check doesn't need Excel to prove its documented outcome. The fourth
    `excel_required` item, `power_query_minimal`, carries a genuine
    (not placeholder) connection and IS pushed through the real supervisor.
    """

    name: str
    input_path: Path
    intent: str
    expected_backend: str
    description: str
    exercise_supervisor: bool = False
    steps: tuple = ()
    contract: dict | None = None
    expected_contract_pass: bool | None = None
    expected_supervisor_ok: bool = True
    """Whether the supervisor job is expected to report ok=True. Default True
    (the ordinary case). power_query_minimal sets this False: on this
    machine, EXCEL.EXE reproducibly does not exit within any tested budget
    (60s, then 300s) after a genuine Power Query refresh, even though the
    actual work (refresh/calc/save) completes in seconds and produces a
    correct output file -- see its own description and README.md. The
    supervisor's TIMED_OUT verdict in that case is the CORRECT, conservative
    behavior (it cannot confirm clean process teardown, so it does not claim
    success), not a bug -- but a caller of this harness needs a way to say
    "TIMED_OUT is the expected, documented outcome here," which is what this
    field is for."""
    macro_check: tuple | None = None  # (macro_name, allowlist) for control_plane.macro_policy


def _build_plain_workbook(path: Path) -> None:
    """A plain xlsx: formulas, no macros/signature/links/data model/pivots/
    slicers/embedded objects. Expected router decision: openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["ID", "Quantity", "Price"])
    ws.append([1, 10, 2.5])
    ws.append([2, 5, 4.0])
    ws.append([3, 8, 1.25])
    ws["E2"] = "=B2*C2"
    ws["E3"] = "=B3*C3"
    ws["E4"] = "=B4*C4"
    ws["G1"] = _SENTINEL_MARKER
    wb.save(path)


def _build_macro_workbook(path: Path) -> None:
    """An .xlsm with a placeholder xl/vbaProject.bin zip entry.

    Not a real compiled VBA project -- fabricating genuinely compiled VBA
    bytecode from scratch, without Excel or a third-party OLE-writer
    library, is out of scope for this issue. workbook_inventory.py's
    has_macros detection is an exact zip-namelist match on
    "xl/vbaProject.bin" (see its module docstring: it inspects the raw
    package, never opens it with openpyxl or Excel), so a placeholder blob
    at that exact path is sufficient to trigger the router decision this
    item exists to prove. This file's content types are not adjusted to
    declare a macro-enabled workbook, and it is never opened by Excel COM
    in this issue's harness -- see run_corpus.py, which checks macro
    rejection at the control_plane.macro_policy layer instead of by
    actually invoking the supervisor's (unimplemented, per #36's own
    README) run_approved_macro step.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Macro-enabled corpus fixture (placeholder VBA project; not executable)."
    wb.save(path)

    with zipfile.ZipFile(path, "a", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("xl/vbaProject.bin", b"placeholder-not-a-real-compiled-vba-project")


def _build_table_workbook(path: Path) -> None:
    """A plain xlsx with one worksheet Table (ListObject) holding two data
    rows. No workbook connection -- a real live Power-Query-backed
    connection refresh is already proven by the C# integration test
    PerConnectionRefreshTests (#36).

    This exact connection-free shape is deliberately shared by two
    independent callers, and must stay connection-free for both:
    `benchmark.py` calls this function directly and relies on it producing
    zero connections (see its own "Benchmark scope deviation" docstring/
    README section). `build_corpus()` below also calls this function for
    the `table_connection` corpus item's base shape, then separately layers
    a placeholder `xl/connections.xml` part onto the file it wrote -- as a
    build_corpus()-local step, not inside this function -- specifically to
    prove the issue #70 `has_connections` routing fix. Do not add a
    connection here; add it at that call site instead."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales"
    ws.append(["Region", "Amount"])
    ws.append(["East", 100])
    ws.append(["West", 200])

    table = Table(displayName="SalesData", ref="A1:B3")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(table)

    wb.save(path)


def _build_external_link_workbook(path: Path) -> None:
    """A plain xlsx plus one inert, unreferenced xl/externalLinks/*.xml zip
    entry. Expected router decision: excel_required.

    The injected part is not wired into any real relationship/content-type
    declaration -- it exists purely so workbook_inventory.py's prefix match
    on "xl/externallinks/" fires, exactly like the macro item above.
    Deliberately never exercised through the real supervisor in
    run_corpus.py (this item's only documented expected outcome is the
    router decision) -- there is no reason to find out how real Excel's
    repair/recovery UI would react to a placeholder external-link part on a
    machine the owner actively uses, and the issue's own corpus item
    description for this entry makes no claim beyond the router decision.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Workbook with an inert, placeholder external-workbook-link OOXML part."
    wb.save(path)

    with zipfile.ZipFile(path, "a", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(
            "xl/externalLinks/externalLink1.xml",
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            b'<externalLink xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"/>',
        )


def _build_failing_contract_workbook(path: Path) -> None:
    """A plain, otherwise-unremarkable xlsx paired (by run_corpus.py) with a
    validation contract designed to fail -- proving the certification
    harness actually detects and reports a failure, not just happy paths."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["ID", "Value"])
    ws.append([1, 100])
    wb.save(path)


def _run_power_query_fixture_script(
    workbook_path: Path,
    query_name: str,
    mformula_path: Path,
    worksheet_name: str,
) -> None:
    """Invoke fixtures/build_power_query_fixture.ps1 as a subprocess to build
    one genuine query + worksheet load. Launches Excel. Caller must have
    already satisfied excel_safety's preflight gate."""
    args = [
        "powershell",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(_POWER_QUERY_FIXTURE_SCRIPT),
        "-WorkbookPath",
        str(workbook_path),
        "-QueryName",
        query_name,
        "-MFormulaPath",
        str(mformula_path),
        "-WorksheetName",
        worksheet_name,
    ]

    result = subprocess.run(args, capture_output=True, text=True, timeout=180, check=False)
    if result.returncode != 0:
        raise RuntimeError(
            f"build_power_query_fixture.ps1 failed (exit {result.returncode}): "
            f"stdout={result.stdout!r} stderr={result.stderr!r}"
        )


def build_power_query_item(directory: str | Path) -> CorpusItem:
    """Build the one corpus item with a genuine Power Query M connection, by
    shelling out to fixtures/build_power_query_fixture.ps1 (see module
    docstring). Launches Excel. Caller MUST have already satisfied
    excel_safety's preflight gate -- this function performs no safety check
    of its own and is not called by build_corpus().

    Expected router decision: excel_required (FIXED by issue #70). This used
    to be the SAME known gap as table_connection above (workbook connections
    were not one of the seven risk fields workbook_inventory.py tracked) --
    and a more consequential instance of it: SKILL.md's own existing
    guidance is explicit that Power Query M work must go through Excel COM,
    never file-only libraries, so routing a genuinely M-code-backed
    workbook to openpyxl for an "edit" was not merely a suboptimal choice
    but contrary to the skill's own documented rule. Confirmed by direct
    measurement against this exact workbook, not assumed. Now that
    workbook_inventory.py detects xl/connections.xml (has_connections) and
    file_router.py treats it as a risk field, this item's router-decision
    check proves the fix instead of documenting the gap.
    """
    directory = Path(directory)
    directory.mkdir(parents=True, exist_ok=True)

    path = directory / "power_query_minimal.xlsx"
    # build_power_query_fixture.ps1 opens an existing workbook
    # (Workbooks.Open) -- it does not create one -- so a blank workbook must
    # exist first.
    Workbook().save(path)

    m_formula = (
        "let\n"
        "    Source = Table.FromRows(\n"
        '        {{1, "Alpha"}, {2, "Beta"}, {3, "Gamma"}},\n'
        '        {"Id", "Name"}\n'
        "    )\n"
        "in\n"
        "    Source"
    )
    m_path = directory / "power_query_minimal.m"
    m_path.write_text(m_formula, encoding="utf-8")

    _run_power_query_fixture_script(path, "SyntheticSource", m_path, "Loaded")

    return CorpusItem(
        name="power_query_minimal",
        input_path=path,
        intent="edit_existing",
        expected_backend="excel_required",
        description=(
            "Workbook with a genuine, self-contained Power Query M connection "
            "(Table.FromRows -- no external file or network dependency), built via the "
            "test-fixture-only build_power_query_fixture.ps1 rather than a hand-crafted "
            "placeholder part. "
            "Expected router decision: excel_required -- FIXED (issue #70): this used to be "
            "the same known gap as table_connection (workbook connections were not a "
            "tracked risk field), confirmed against a genuine M-code-backed connection "
            "rather than only a plain Table; now that workbook_inventory.py's has_connections "
            "detects xl/connections.xml and file_router.py fails closed on it, this item's "
            "router-decision check proves the fix rather than documenting the gap.\n\n"
            "MAJOR FINDING, then FIXED and reverified: the first two real runs against this "
            "item (60s and 300s refresh/close budgets) reproducibly hit a genuine defect -- "
            "the job's actual work (open, refresh, recalc, save) completed in ~9 seconds "
            "every time (confirmed via events.jsonl reaching SAVING then SUCCEEDED, and a "
            "correct output.xlsx), but EXCEL.EXE itself never exited within either budget, "
            "so the supervisor correctly timed out and force-killed via the Job Object "
            "(zero orphaned processes both times). Root cause, found via web research and "
            "confirmed by code inspection, not guessed: StepRunner.cs's RunRefresh obtained "
            "Workbook.Connections, each Connection, and each OLEDBConnection/ODBCConnection "
            "via COM property access and never released any of them -- classic unreleased-"
            "RCW behavior (Application.Quit() only requests an exit; Excel won't actually "
            "terminate until every outstanding COM reference is released, and relying on the "
            ".NET GC alone to get there can take multiple collection cycles, or apparently "
            "never resolve within any tested budget for a genuine Power-Query/Mashup-backed "
            "connection specifically). Fixed by explicit Marshal.ReleaseComObject on every "
            "one of those intermediate objects (in finally blocks, surviving exceptions) "
            "plus strengthening ExcelSession.cs's single GC.Collect()+WaitForPendingFinalizers "
            "to the standard double-collect pattern. Reverified against this exact item after "
            "the fix: SUCCEEDED, ok=True, in 12.5 seconds -- down from timing out past 300s. "
            "See README.md for the full writeup and sources."
        ),
        exercise_supervisor=True,
        steps=(
            {"type": "open", "workbook_path": str(path)},
            {"type": "refresh", "connections": "all"},
            {"type": "recalc", "mode": "full_rebuild"},
        ),
        contract={
            "required_sheets": ["Loaded"],
            "min_row_counts": {"Loaded": 4},
            "sentinel_cells": [{"sheet": "Loaded", "cell": "B2", "expected": "Alpha"}],
        },
        expected_contract_pass=True,
    )


def build_corpus(directory: str | Path) -> list[CorpusItem]:
    """Generate every corpus workbook into `directory` (created if needed)
    and return the list of CorpusItem descriptors, each carrying its own
    documented expected outcome. Never touches Excel."""
    directory = Path(directory)
    directory.mkdir(parents=True, exist_ok=True)

    plain_path = directory / "plain_formulas.xlsx"
    _build_plain_workbook(plain_path)

    macro_path = directory / "macro_enabled.xlsm"
    _build_macro_workbook(macro_path)

    table_path = directory / "table_connection.xlsx"
    _build_table_workbook(table_path)
    # Issue #70: layer a placeholder, inert xl/connections.xml part onto the
    # table workbook -- same pattern as the macro/external-link placeholders
    # above (not wired into any real relationship/content-type declaration;
    # workbook_inventory.py's has_connections detection is an exact
    # zip-namelist match, so a placeholder at the exact path is sufficient
    # to prove this item's router decision). Added here, not inside
    # _build_table_workbook itself, so benchmark.py's own direct call to
    # that function stays genuinely connection-free (see that function's
    # docstring). Because this part is a placeholder, not a real connection
    # declaration, this item is router-decision-only below -- never opened
    # by real Excel in this harness (see the "table_connection" CorpusItem).
    with zipfile.ZipFile(table_path, "a", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(
            "xl/connections.xml",
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
            b'<connections xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"/>',
        )

    external_link_path = directory / "external_link.xlsx"
    _build_external_link_workbook(external_link_path)

    failing_contract_path = directory / "failing_contract.xlsx"
    _build_failing_contract_workbook(failing_contract_path)

    return [
        CorpusItem(
            name="plain_formulas",
            input_path=plain_path,
            intent="edit_existing",
            expected_backend="openpyxl",
            description=(
                "Plain workbook, formulas, no risk features. Expected router decision: "
                "openpyxl. Should pass a simple validation contract evaluated directly "
                "against this file (no Excel involved)."
            ),
            exercise_supervisor=False,
            contract={
                "required_sheets": ["Data"],
                "min_row_counts": {"Data": 4},
                "sentinel_cells": [{"sheet": "Data", "cell": "G1", "expected": _SENTINEL_MARKER}],
            },
            expected_contract_pass=True,
        ),
        CorpusItem(
            name="macro_enabled",
            input_path=macro_path,
            intent="edit_existing",
            expected_backend="excel_required",
            description=(
                "Macro-enabled workbook (.xlsm with a placeholder xl/vbaProject.bin part). "
                "Expected router decision: excel_required. This item's macro_check only "
                "proves two narrower things, and does NOT prove Excel-level macro "
                "rejection or that a genuinely macro-bearing workbook is treated "
                "differently from a plain one at execution time: (1) "
                "workbook_inventory.inspect_workbook(...).has_macros is True for this "
                "file (the router's zip-namelist detection actually fired), and (2) "
                "macro_policy.is_macro_approved(...) -- a pure sha256+entrypoint lookup "
                "against a caller-supplied allowlist, with no awareness of the "
                "workbook's actual content -- returns False against an empty allowlist, "
                "which it would do identically for any input file, macro-bearing or not. "
                "The supervisor's own run_approved_macro step (the one place that would "
                "actually gate execution at the Excel/COM level) is unimplemented -- see "
                "supervisor/README.md ('MACRO_EXECUTION_DEFERRED') -- and this corpus item "
                "does not exercise it. 'Prove Excel-level AutomationSecurity macro "
                "rejection' remains an open gap, tracked here and in README.md, not "
                "something this check stands in for."
            ),
            exercise_supervisor=False,
            macro_check=("TestMacro", []),
        ),
        CorpusItem(
            name="table_connection",
            input_path=table_path,
            intent="edit_existing",
            expected_backend="excel_required",
            description=(
                "Worksheet Table with two data rows, plus a placeholder xl/connections.xml "
                "part (not wired into any real relationship/content-type declaration -- see "
                "build_corpus()). FIXED (issue #70): this item used to document a real gap -- "
                "file_router.py had no tracked field for 'has a workbook connection', so a "
                "connection-bearing workbook routed to openpyxl instead of excel_required. "
                "Now that workbook_inventory.py detects xl/connections.xml and file_router.py "
                "treats has_connections as a risk field, this item's router-decision check "
                "proves the fix: expected router decision is excel_required. Router-decision-"
                "only -- because the injected part is a placeholder, not a genuine connection "
                "declaration, this item is deliberately never pushed through real Excel in "
                "this harness (same reasoning as the macro_enabled/external_link items). The "
                "original 'drive a refresh step end-to-end against a Table-bearing workbook "
                "through the real supervisor' purpose this item used to serve is still "
                "covered, unaffected, by benchmark.py's own direct use of the underlying "
                "connection-free _build_table_workbook shape (both its supervisor and legacy "
                "legs)."
            ),
            exercise_supervisor=False,
        ),
        CorpusItem(
            name="external_link",
            input_path=external_link_path,
            intent="edit_existing",
            expected_backend="excel_required",
            description=(
                "Workbook with an external workbook link (placeholder OOXML part). "
                "Expected router decision: excel_required. Router-decision-only -- never "
                "pushed through real Excel in this harness (see this module's build "
                "function for why)."
            ),
            exercise_supervisor=False,
        ),
        CorpusItem(
            name="failing_contract",
            input_path=failing_contract_path,
            intent="edit_existing",
            expected_backend="openpyxl",
            description=(
                "Otherwise-plain workbook paired with a validation contract designed to "
                "fail (a row-count minimum far above the actual data), to prove the "
                "certification harness actually detects and reports a failure rather "
                "than only exercising happy paths."
            ),
            exercise_supervisor=False,
            contract={
                "required_sheets": ["Data"],
                "min_row_counts": {"Data": 50},
            },
            expected_contract_pass=False,
        ),
    ]
